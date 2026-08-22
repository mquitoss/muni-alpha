"""Official source locations and source-specific parsers."""

from __future__ import annotations

import json
import math
import re
import xml.etree.ElementTree as ET
from html import unescape
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

from openpyxl import load_workbook

ICGC_WFS = "https://geoserveis.icgc.cat/servei/catalunya/divisions-administratives/wfs"
MUNICIPALITIES_LAYER = "divisions_administratives_wfs:divisions_administratives_municipis_5000"
CAPITALS_LAYER = "divisions_administratives_wfs:divisions_administratives_capsdemunicipi_capmunicipi"
SALES_BASE_URL = (
    "https://habitatge.gencat.cat/web/.content/home/dades/estadistiques/"
    "01_Estadistiques_de_construccio_i_mercat_immobiliari/"
    "02_Compravenda_i_preu_de_venda/"
    "02_Compravendes_d_habitatges_registrades_i_el_preu_de_venda"
)
RENTALS_API = "https://analisi.transparenciacatalunya.cat/resource/qww9-bvhh.json"
# The table API currently limits this request to six periods and answers HTTP
# 416 instead of clipping larger ranges.
POPULATION_URL = "https://api.idescat.cat/taules/v2/ep/9122/10002/mun/data?_LAST_=6"
ANNUAL_POPULATION_URL = (
    "https://api.idescat.cat/taules/v2/pmh/446/477/mun/data?SEX=TOTAL&_LAST_=6"
)
INCOME_URL = "https://api.idescat.cat/taules/v2/rfdbc/21181/25017/mun/data?_LAST_=7"
ETCA_URL = "https://www.idescat.cat/pub/?id=epe&n=17886&geo=mun&lang=ca&f=ssv"
TOURISM_REGISTRY_API = "https://analisi.transparenciacatalunya.cat/resource/t2h3-cgys.json"
HUT_REGULATION_URL = "https://portaljuridic.gencat.cat/eli/es-ct/dl/2023/11/07/3/dof/cat/xml"


def wfs_url(layer: str, *, properties: str | None = None) -> str:
    parameters = {
        "service": "WFS",
        "version": "2.0.0",
        "request": "GetFeature",
        "typeNames": layer,
        "outputFormat": "GEOJSON",
        "srsName": "EPSG:4326",
        "count": "2000",
    }
    if properties:
        parameters["propertyName"] = properties
    return f"{ICGC_WFS}?{urlencode(parameters)}"


def rentals_url(year: int = 2025) -> str:
    query = {
        "$limit": "50000",
        "$where": f"any='{year}' AND periode='gener-desembre'",
        "$order": "codi_territorial",
    }
    return f"{RENTALS_API}?{urlencode(query)}"


def sales_url(year: int) -> str:
    return f"{SALES_BASE_URL}/{year}/MUN_acum1any_{year}.xlsx"


def tourism_registry_url() -> str:
    query = {
        "$select": (
            "codi_municipi_idescat,tipus_establiment,"
            "count(distinct n_mero_inscripci) as n"
        ),
        "$where": "estat='Alta'",
        "$group": "codi_municipi_idescat,tipus_establiment",
        "$limit": "50000",
    }
    return f"{TOURISM_REGISTRY_API}?{urlencode(query)}"


def parse_municipalities(catalogue_path: Path, capitals_path: Path) -> list[dict[str, Any]]:
    # ICGC emits `{"type":"MultiPolygon",}` when geometry was excluded via
    # propertyName. Repair that server-side serialization defect locally.
    catalogue_text = catalogue_path.read_text(encoding="utf-8")
    catalogue = json.loads(re.sub(r",\s*}", "}", catalogue_text))
    capitals = json.loads(capitals_path.read_text(encoding="utf-8"))
    capital_by_code = {
        feature["properties"]["CODIMUNI"]: feature.get("geometry", {}).get("coordinates", [])
        for feature in capitals["features"]
    }
    rows = []
    for feature in catalogue["features"]:
        item = feature["properties"]
        coordinates = capital_by_code.get(item["CODIMUNI"], [])
        rows.append(
            {
                "municipality_code": item["CODIMUNI"],
                "municipality_name": item["NOMMUNI"],
                "comarca_code": item["CODICOMAR"],
                "comarca_name": item["NOMCOMAR"],
                "province_code": item["CODIPROV"],
                "province_name": item["NOMPROV"],
                "territorial_area_code": item["CODIVEGUE"],
                "territorial_area_name": item["NOMVEGUE"],
                "capital_lat": coordinates[1] if len(coordinates) == 2 else None,
                "capital_lon": coordinates[0] if len(coordinates) == 2 else None,
                "area_km2": item["AREAM5000"],
                "geometry_source_date": None,
            }
        )
    rows.sort(key=lambda row: row["municipality_code"])
    _validate_codes(rows)
    return rows


def parse_sales(path: Path, year: int = 2025) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = f"4t{str(year)[-2:]}acum_1any"
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"expected worksheet {sheet_name!r}; got {workbook.sheetnames!r}")
    rows: list[dict[str, Any]] = []
    for cells in workbook[sheet_name].iter_rows(min_row=1, values_only=True):
        code = str(cells[0]).strip() if cells[0] is not None else ""
        if not (len(code) == 5 and code.isdigit()):
            continue
        rows.append(
            {
                "ine_code": code,
                "source_name": cells[1],
                "sales_count": _number(cells[4]),
                "avg_surface_m2": _number(cells[9]),
                "avg_sale_price_eur": _scaled_number(cells[13], 1000),
                "sale_price_eur_m2": _number(cells[17]),
            }
        )
    return rows


def parse_rentals(path: Path) -> list[dict[str, Any]]:
    source = json.loads(path.read_text(encoding="utf-8"))
    return [
        {
            "ine_code": str(item.get("codi_territorial", "")).zfill(5),
            "source_name": item.get("nom_territori"),
            "rental_contracts": _number(item.get("habitatges")),
            "avg_monthly_rent_eur": _number(item.get("renda")),
            "reference_period": item.get("any"),
        }
        for item in source
        if str(item.get("codi_territorial", "")).isdigit()
    ]


def parse_population(path: Path) -> dict[str, dict[str, float]]:
    dataset = json.loads(path.read_text(encoding="utf-8"))
    ids = dataset["id"]
    sizes = dataset["size"]
    dimensions = dataset["dimension"]
    labels = {dimension: _dimension_codes(dimensions[dimension]) for dimension in ids}
    values = dataset["value"]
    result: dict[str, dict[str, float]] = {}
    for semester_index, semester in enumerate(labels["SEMESTER"]):
        for municipality_index, municipality in enumerate(labels["MUN"]):
            if municipality == "TOTAL":
                continue
            coordinates = {
                "SEMESTER": semester_index,
                "MUN": municipality_index,
                "SEX": labels["SEX"].index("TOTAL"),
                "CONCEPT": labels["CONCEPT"].index("POP"),
            }
            flat_index = 0
            for dimension, size in zip(ids, sizes, strict=True):
                flat_index = flat_index * size + coordinates[dimension]
            value = values.get(str(flat_index)) if isinstance(values, dict) else values[flat_index]
            if value is not None:
                result.setdefault(municipality, {})[semester] = float(value)
    return result


def parse_annual_population(path: Path) -> dict[str, dict[str, float]]:
    dataset = json.loads(path.read_text(encoding="utf-8"))
    if dataset.get("id") != ["YEAR", "MUN", "SEX", "CONCEPT"]:
        raise ValueError("unexpected annual population dimensions")
    dimensions = dataset["dimension"]
    years = _dimension_codes(dimensions["YEAR"])
    municipalities = _dimension_codes(dimensions["MUN"])
    if _dimension_codes(dimensions["SEX"]) != ["TOTAL"]:
        raise ValueError("annual population source is not filtered to total sex")
    if _dimension_codes(dimensions["CONCEPT"]) != ["POP"]:
        raise ValueError("annual population source has an unexpected concept")
    values = dataset["value"]
    result: dict[str, dict[str, float]] = {}
    for year_index, year in enumerate(years):
        for municipality_index, municipality in enumerate(municipalities):
            if municipality == "TOTAL":
                continue
            flat_index = year_index * len(municipalities) + municipality_index
            value = values.get(str(flat_index)) if isinstance(values, dict) else values[flat_index]
            if value is not None:
                result.setdefault(municipality, {})[year] = float(value)
    return result


def parse_income(path: Path) -> tuple[str, dict[str, dict[str, float]]]:
    dataset = json.loads(path.read_text(encoding="utf-8"))
    ids = dataset["id"]
    sizes = dataset["size"]
    dimensions = dataset["dimension"]
    labels = {dimension: _dimension_codes(dimensions[dimension]) for dimension in ids}
    latest_year = labels["YEAR"][-1]
    values = dataset["value"]
    result: dict[str, dict[str, float]] = {}
    for municipality_index, municipality in enumerate(labels["MUN"]):
        if municipality == "TOTAL":
            continue
        for indicator_index, indicator in enumerate(labels["INDICATOR"]):
            coordinates = {
                "YEAR": labels["YEAR"].index(latest_year),
                "MUN": municipality_index,
                "CONCEPT": 0,
                "INDICATOR": indicator_index,
            }
            flat_index = 0
            for dimension, size in zip(ids, sizes, strict=True):
                flat_index = flat_index * size + coordinates[dimension]
            value = values.get(str(flat_index)) if isinstance(values, dict) else values[flat_index]
            if value is not None:
                result.setdefault(municipality, {})[indicator] = float(value)
    return latest_year, result


def parse_etca(path: Path) -> tuple[str, dict[str, dict[str, float]]]:
    import csv
    import io

    text = path.read_text(encoding="utf-8-sig")
    lines = text.splitlines()
    period_line = next((line for line in lines if line.startswith("Municipis. ")), "")
    period = period_line.removeprefix("Municipis. ").split()[0]
    header_index = next(
        (index for index, line in enumerate(lines) if line.startswith("Codi;Nom;")),
        None,
    )
    if header_index is None:
        raise ValueError("ETCA source does not contain the expected header")
    reader = csv.DictReader(io.StringIO("\n".join(lines[header_index:])), delimiter=";")
    result: dict[str, dict[str, float]] = {}
    for row in reader:
        code = (row.get("Codi") or "").strip()
        if len(code) != 6 or not code.isdigit():
            continue
        values = list(row.values())
        result[code] = {
            "non_resident_present": _number(values[2]) or 0,
            "resident_absent": _number(values[3]) or 0,
            "seasonal_population_total": _number(values[4]) or 0,
            "resident_population": _number(values[5]) or 0,
            "population_etca": _number(values[6]) or 0,
            "population_etca_pct": _number(values[7]) or 0,
        }
    return period, result


def parse_tourism_registry(path: Path) -> dict[str, dict[str, int]]:
    source = json.loads(path.read_text(encoding="utf-8"))
    categories = {
        "habitatges d'ús turístic": "hut_count",
        "habitatge d'ús turístic": "hut_count",
        "hotels": "hotel_count",
        "establiment hoteler": "hotel_count",
        "turisme rural": "rural_count",
        "càmpings": "camping_count",
        "càmping": "camping_count",
    }
    result: dict[str, dict[str, int]] = {}
    for item in source:
        code = str(item.get("codi_municipi_idescat", ""))
        category = categories.get(str(item.get("tipus_establiment", "")).strip().casefold())
        if len(code) != 6 or not code.isdigit() or category is None:
            continue
        counts = result.setdefault(
            code,
            {"hut_count": 0, "hotel_count": 0, "rural_count": 0, "camping_count": 0},
        )
        counts[category] += int(item.get("n", 0))
    return result


class _AnnexTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.in_cell = False
        self.cell_parts: list[str] = []
        self.row: list[str] = []
        self.names: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "td":
            self.in_cell = True
            self.cell_parts = []

    def handle_data(self, data: str) -> None:
        if self.in_cell:
            self.cell_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag == "td" and self.in_cell:
            self.row.append(" ".join("".join(self.cell_parts).split()))
            self.in_cell = False
        elif tag == "tr":
            if len(self.row) >= 2 and self.row[0].isdigit():
                self.names.append(self.row[1])
            self.row = []


def parse_hut_regime(
    path: Path, municipalities: list[dict[str, Any]]
) -> set[str]:
    root = ET.parse(path).getroot()
    table_html = None
    for container in root.iter():
        if container.tag.rsplit("}", 1)[-1] != "hcontainer":
            continue
        heading = next(
            (
                child
                for child in container
                if child.tag.rsplit("}", 1)[-1] == "heading"
            ),
            None,
        )
        heading_text = "".join(heading.itertext()) if heading is not None else ""
        if "Relació de municipis" not in heading_text:
            continue
        content = next(
            (
                child
                for child in container
                if child.tag.rsplit("}", 1)[-1] == "content"
            ),
            None,
        )
        if content is not None:
            table_html = unescape(content.attrib.get("period", ""))
            break
    if not table_html:
        raise ValueError("HUT regulation annex was not found in the official XML")

    parser = _AnnexTableParser()
    parser.feed(table_html)
    if len(parser.names) != 262 or len(set(parser.names)) != 262:
        raise ValueError(f"expected 262 unique HUT annex municipalities, got {len(parser.names)}")

    code_by_name = {row["municipality_name"]: row["municipality_code"] for row in municipalities}
    aliases = {
        "Castell-Platja d'Aro": "Castell d'Aro, Platja d'Aro i s'Agaró",
    }
    codes = set()
    unresolved = []
    for name in parser.names:
        canonical_name = aliases.get(name, name)
        code = code_by_name.get(canonical_name)
        if code is None:
            unresolved.append(name)
        else:
            codes.add(code)
    if unresolved or len(codes) != 262:
        raise ValueError(f"unresolved HUT annex municipalities: {unresolved!r}")
    return codes


def _dimension_codes(dimension: dict[str, Any]) -> list[str]:
    index = dimension["category"]["index"]
    if isinstance(index, list):
        return index
    return [code for code, _ in sorted(index.items(), key=lambda item: item[1])]


def _number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, str):
        cleaned = value.strip().lower().replace("€", "").replace(" ", "")
        if cleaned in {"", "n.d.", "nd", "-"}:
            return None
        cleaned = cleaned.replace(",", ".")
    else:
        cleaned = value
    try:
        number = float(cleaned)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _scaled_number(value: Any, scale: float) -> float | None:
    number = _number(value)
    return number * scale if number is not None else None


def _validate_codes(rows: list[dict[str, Any]]) -> None:
    codes = [row["municipality_code"] for row in rows]
    if any(len(code) != 6 or not code.isdigit() for code in codes):
        raise ValueError("ICGC returned an invalid municipality code")
    if len(codes) != len(set(codes)):
        raise ValueError("ICGC returned duplicate municipality codes")
